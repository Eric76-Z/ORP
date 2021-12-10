# 获取机器人数据
import json
import os
import re
import time
import zipfile
from difflib import SequenceMatcher

from tqdm import tqdm
from openpyxl import Workbook, load_workbook

from src.common.rob_program_backup_setting import SH_LOG_TITLE
from src.common.rob_program_comment_setting import RobProgramComment, SH_ROB_COMMENT_ANALYSE, PATH_PROGRAM, \
    ROB_COMMENT_OVERVIEW_REPORT_BASE, ROB_COMMENT_OVERVIEW_REPORT, \
    SH_ROB_COMMENT_OVERVIEW, STANDARD_COMMENT, BID_DATA_TEMPLE, SH_BID_DATA_TEMPLE, SUM
from src.common.setting import PATH_BASE, time_now, GET_BIG_DATA_SIMPLE
from src.common.utils import TimeStampToTime, getFileSize, logWrite, createFolder, createSheet, backSignalSort, \
    process_bar


# 1、读取机器人zip相关信息
def RobotInfo(standard_comment, GET_BIG_DATA_SIMPLE):
    print('========================overview表格--start========================')
    createFolder(ROB_COMMENT_OVERVIEW_REPORT_BASE)
    wb = Workbook(write_only=True)
    createSheet(wb=wb, sh_name='机器人注释解析总览', sh_index=1, sh_title=SH_ROB_COMMENT_OVERVIEW)
    createSheet(wb=wb, sh_name='log', sh_index=2, sh_title=SH_LOG_TITLE)
    ov_path = os.path.join(ROB_COMMENT_OVERVIEW_REPORT_BASE, ROB_COMMENT_OVERVIEW_REPORT)
    wb.save(ov_path)
    wb.close()
    wb = load_workbook(ov_path)

    rob_program_comment_json = {}
    big_data_temple = {}
    i = 0
    for root, dirs, files in os.walk(PATH_PROGRAM):
        # 显示进度条
        for name in tqdm(files):
            originpath = os.path.join(root, name)
            if zipfile.is_zipfile(originpath):
                rob_program_comment = RobProgramComment()
                # ================path==================
                rob_program_comment.path['path_origin'] = originpath  # 原始路径
                # ================meta==================
                rob_program_comment.meta['title'] = name  # 文件名
                rob_program_comment.meta['format'] = 'zip'  # 文件后缀
                rob_program_comment.meta['mtime'] = TimeStampToTime(os.path.getmtime(originpath))  # 原数据中修改时间，可视为创建时间
                rob_program_comment.meta['size'] = getFileSize(originpath)
                # ================data==================
                rob_program_comment.data['filename'] = name
                rob_program_comment.data['controllername'] = name.split('.zip')[0]  # eg.k2a3a131s460r04
                rob_program_comment.data['workstationname'] = rob_program_comment.data['controllername'][
                                                              -7:].upper()  # 截取的 eg.s460r04
                rob_program_comment.data['localLv1'] = rob_program_comment.localLv1()
                rob_program_comment.data['localLv2'] = rob_program_comment.localLv2()
                rob_program_comment.data['localLv3'] = rob_program_comment.localLv3()

                if GET_BIG_DATA_SIMPLE == True:
                    big_data_temple = backBigDataSimple(rob_program_comment, big_data_temple)
                else:
                    # 创建文件夹
                    ROB_COMMENT_ANALYSE_REPORT_BASE = PATH_BASE + '\\' + '机器人注释解析详情报告' + '\\' + str(
                        time_now.year) + str(
                        time_now.month) + str(
                        time_now.day) + '\\' + str(
                        time_now.hour) + str(
                        time_now.minute) + str(time_now.second) + '\\' + rob_program_comment.data[
                                                          'localLv1'] + '\\' + \
                                                      rob_program_comment.data['localLv2'] + '\\' + \
                                                      rob_program_comment.data[
                                                          'localLv3']
                    # ROB_COMMENT_ANALYSE_REPORT_TIME =
                    createFolder(ROB_COMMENT_ANALYSE_REPORT_BASE)
                    ROB_COMMENT_ANALYSE_REPORT = rob_program_comment.data['controllername'] + '.xlsx'
                    ay_path = os.path.join(ROB_COMMENT_ANALYSE_REPORT_BASE, ROB_COMMENT_ANALYSE_REPORT)  # 保存路径
                    # print('========================解析机器人备份--start========================')
                    analysisZip(rob_program_comment, wb=wb, standard_comment=standard_comment, path=ay_path)
                    # print('========================解析机器人备份--end========================')
                    # 机器人overview表格写入
                    ws = wb['机器人注释解析总览']
                    i = i + 1
                    ws.append([i, 'PFH2B', rob_program_comment.data['localLv1'],
                               rob_program_comment.data['localLv2'],
                               rob_program_comment.data['localLv3'], rob_program_comment.data['workstationname'],
                               rob_program_comment.meta['mtime']])
                    ay_path = ay_path.replace(PATH_BASE, '..\\..')
                    ws.cell(row=i + 1, column=8).hyperlink = ay_path
                    # 添加入rob_program_comment_json
                    rob_program_comment_json[rob_program_comment.data['workstationname']] = rob_program_comment

            else:
                msg = '备份可能损坏!!!源路径为：' + originpath + ';'
                logWrite(wb=wb, controllername=name, sort='警告', msg=msg)
    if GET_BIG_DATA_SIMPLE == False:
        # 写入json文件
        with open('database/robot_comment.json', "w", encoding='utf-8') as f:
            f.write(json.dumps(rob_program_comment_json, default=lambda obj: obj.__dict__, indent=4,
                               ensure_ascii=False))
            f.close()
        wb.save(ov_path)
        # wb.save(PATH_BASE + '\\' + '机器人注释报告总览new.xlsx')
        wb.close()
        print('========================overview表格--end========================')
    else:
        # 写入json
        with open('database/recommend.json', "w", encoding='utf-8') as f:
            f.write(
                json.dumps(big_data_temple, default=lambda obj: obj.__dict__, indent=4, ensure_ascii=False))
            f.close()
            # 创建工作表，写入内容
            wb4 = Workbook(write_only=True)
            createSheet(wb=wb4, sh_name='机器人注释大数据统计', sh_index=1,
                        sh_title=SH_BID_DATA_TEMPLE)
            wb4.save(BID_DATA_TEMPLE)
            wb4.close()

            wb4_reopen = load_workbook(BID_DATA_TEMPLE)
            ws4_reopen = wb4_reopen['机器人注释大数据统计']
            i = 0
            for b in big_data_temple:
                i = i + 1
                # print(big_data_temple[b]['recommend'])
                key, = big_data_temple[b]['recommend']
                # print(key)
                list = [i, b, '总数' + str(big_data_temple[b]['sum']), key, big_data_temple[b]['recommend'][key]['num'],
                        big_data_temple[b]['recommend'][key]['ratio']]
                for c in big_data_temple[b]['comments']:
                    list = list + [c, big_data_temple[b]['comments'][c]['num'],
                                   big_data_temple[b]['comments'][c]['ratio']]
                ws4_reopen.append(list)
            wb4_reopen.save(BID_DATA_TEMPLE)
            wb4_reopen.close()


def analysisZip(rob_program_comment, wb, standard_comment, path):
    try:
        filezip = zipfile.ZipFile(rob_program_comment.path['path_origin'], "r")
        rob_program_comment.zipData['total_files'] = len(filezip.namelist())
        try:
            for file in filezip.namelist():
                if file.split('/')[-1] == 'RefListe.txt':
                    create_time = filezip.getinfo(file).date_time
                    rob_program_comment.zipData['create_time'] = str(create_time[0]) + '-' + str(create_time[1]) + '-' + \
                                                                 str(create_time[2]) + ' ' + str(create_time[3]) + ':' + \
                                                                 str(create_time[4]) + ':' + str(create_time[5])
                    RefListe = filezip.open(file)

                    # 创建工作表，写入内容
                    wb2 = Workbook(write_only=True)
                    # 返回的数据
                    ret = analyseRefListe(RefListe)

                    # print(SH_ROB_COMMENT_ANALYSE + ret['used_all'])
                    createSheet(wb=wb2, sh_name='机器人注释解析', sh_index=1,
                                sh_title=(SH_ROB_COMMENT_ANALYSE + ret['used_all']))
                    wb2.save(path)
                    wb2.close()
                    wb2_reopen = load_workbook(path)
                    ws2_reopen = wb2_reopen['机器人注释解析']
                    # # 写入数据
                    i = 0
                    for r in ret['ref_liste']:
                        i = i + 1
                        localLv1 = rob_program_comment.data['localLv1']
                        localLv2 = rob_program_comment.data['localLv2']
                        localLv3 = rob_program_comment.data['localLv3']

                        signal_sort = backSignalSort(r['signal'])
                        # 判断注释是否经过修改
                        s_ratio_max = 0
                        s_comment = standard_comment[r['signal']][0]
                        for sc in standard_comment[r['signal']]:
                            seq = SequenceMatcher(None, r['comments'], sc)
                            ratio = seq.ratio()
                            max_changed = False
                            if ratio > s_ratio_max:
                                s_ratio_max = ratio
                                max_changed = True
                            if s_ratio_max > 0.9 and max_changed == True:
                                s_comment = sc
                        # 根据大数据分析判定此信号的推荐注释
                        recommend = getRecommend(r['signal'])
                        r_ratio = SequenceMatcher(None, r['comments'], recommend).ratio()
                        if s_ratio_max > 0.9 or r_ratio > 0.9:
                            is_changed = 'N'
                        else:
                            is_changed = 'Y'
                        # 判断是否需要注释
                        need_comment = needComment(r['signal'], r['used'])
                        ws2_reopen.append([i, localLv1, localLv2, localLv3, signal_sort, r['signal'],
                                           r['comments'], s_comment, s_ratio_max, recommend, r_ratio, is_changed,
                                           need_comment])
                        j = 0
                        for u in r['used']:
                            j = j + 1
                            index = ret['used_all'].index(u)
                            ws2_reopen.cell(row=i + 1, column=index + 15).value = '*'

                    wb2_reopen.save(path)
                    wb2_reopen.close()
        except Exception as e:
            print(e)
            # print(rob_program_comment.path['path_origin'] + str(e))
        rob_program_comment.zipData['state'] = '备份完好'  # zip文件完好
        filezip.close()
    except Exception as e:
        msg = '备份可能损坏!!!源路径为：' + rob_program_comment.path['path_origin'] + ';' + '新路径为:' + rob_program_comment.path[
            'path_new']
        logWrite(wb=wb, controllername=rob_program_comment.data['controllername'], sort='警告', msg=msg)
        rob_program_comment.zipData['state'] = '备份损坏'
        print(rob_program_comment.zipData['state'])


def backBigDataSimple(rob_program_comment, big_data_temple):
    try:
        filezip = zipfile.ZipFile(rob_program_comment.path['path_origin'], "r")
        rob_program_comment.zipData['total_files'] = len(filezip.namelist())
        try:
            for file in filezip.namelist():
                if file.split('/')[-1] == 'RefListe.txt':
                    RefListe = filezip.open(file)
                    # 返回的数据
                    ref = analyseRefListe(RefListe)
                    for r in ref['ref_liste']:
                        if r['signal'] in big_data_temple:
                            if r['comments'] in big_data_temple[r['signal']]['comments']:
                                # 数量加1
                                big_data_temple[r['signal']]['comments'][r['comments']]['num'] = \
                                    big_data_temple[r['signal']]['comments'][r['comments']]['num'] + 1
                            else:
                                big_data_temple[r['signal']]['comments'][r['comments']] = {'num': 1, 'ratio': 1.0}
                            # 样本总数加1
                            big_data_temple[r['signal']]['sum'] = big_data_temple[r['signal']]['sum'] + 1
                            # 概率
                            temp = 0
                            for c in big_data_temple[r['signal']]['comments']:
                                big_data_temple[r['signal']]['comments'][c]['ratio'] = \
                                    big_data_temple[r['signal']]['comments'][c]['num'] / big_data_temple[r['signal']][
                                        'sum']
                                if big_data_temple[r['signal']]['comments'][c]['num'] > temp:
                                    temp = big_data_temple[r['signal']]['comments'][c]['num']
                                    big_data_temple[r['signal']]['recommend'] = {
                                        c: big_data_temple[r['signal']]['comments'][c]}
                                    # print(big_data_temple[r['signal']]['recommend'])

                            # if big_data_temple[r['signal']]['comments'][r['comments']]['ratio'] != 1.0:
                            #     print(big_data_temple[r['signal']])
                        else:
                            big_data_temple[r['signal']] = {'sum': 1,
                                                            'comments': {r['comments']: {'num': 1, 'ratio': 1.0}},
                                                            'recommend': {r['comments']: {'num': 1, 'ratio': 1.0}}}
            return big_data_temple
        except Exception as e:
            print(e)
    except Exception as e:
        print(e)


# 分析机器人zip中refliste注释文件
def analyseRefListe(RefListe):
    ref_liste = []
    used_all = []
    for l in RefListe.readlines():
        l = str(l)
        l = l.replace('b\'', '')
        l = l.replace('\\r\\n\'', '')
        if re.search(r"^(E|A|M|I|ana|bin|t|F|T|Makro|UP)\s\d{1,4}\s", l) != None:
            # ========================signal========================
            signal = re.search(r"^(E|A|M|I|ana|bin|t|F|T|Makro|UP)\s\d{1,4}", l).group()
            signal = signal.replace(" ", "")  # 去除空格
            # ========================comments========================
            if re.search(r'\[.*]', l) != None:
                if re.search(r'\[\s]', l) != None:
                    comments = 'Reserviert'
                else:
                    comments = re.search(r'\[.*]', l).group()[1:-1]
                    comments = comments.replace(',', '')  # 去除逗号
                    if comments == '':
                        comments = 'Reserviert'
            elif re.search(r'\[]', l) != None:
                comments = 'Reserviert'
            else:
                comments = 'Reserviert'

            # ========================used========================
            if re.search(r']\s*(Folge|Floge|Makro|UP).*$', l) != None:
                used_str = re.search(r']\s*(Folge|Floge|Makro|UP).*$', l).group()
            else:
                used_str = re.search(r'\s\s(Folge|Floge|Makro|UP).*$', l).group()
            # print(used_str)
            used_str = used_str.replace('Floge', 'Folge')
            used_str = used_str.replace('*', '')
            used_str = used_str.replace(']', '')
            used_str = used_str.split('  ')[-1]
            used_str = used_str.replace(' ', '')
            # print(used_str)
            used_list = used_str.split(',')
            used = dealused(used_list)
            ret = {
                'signal': signal,
                'comments': comments,
                'used': used
            }
            ref_liste.append(ret)
        elif re.search(r"^(Folge|Floge)\s\d{1,4}", l) != None:
            signal = 'cell'
            comments = 'Reserviert'
            l = l.replace(' ', '')
            l = l.replace('Floge', 'Folge')
            used_list = l.split(',')

            used = dealused(used_list)
            ret = {
                'signal': signal,
                'comments': comments,
                'used': used
            }
            ref_liste.append(ret)
        else:
            used = []
        used_all = used_all + used
    used_all = list(set(used_all))
    used_all.sort()

    return {
        'ref_liste': ref_liste,
        'used_all': used_all
    }


def getRecommend(signal):
    with open('database/recommend.json', 'r', encoding='utf-8') as f:
        info_dict = json.load(f)
        for dict in info_dict:
            # print(dict)
            if dict == signal:
                key, = info_dict[dict]['recommend']
                return key


def needComment(signal, used):
    if signal.startswith('E'):
        ret = analysisUsed(used)
        return ret
    elif signal.startswith('A'):
        ret = analysisUsed(used)
        return ret
    elif signal.startswith('M'):
        if signal.startswith('Makro'):
            ret = analysisUsed(used)
            return ret
        else:
            ret = analysisUsed(used)
            return ret
    elif signal.startswith('I'):
        ret = analysisUsed(used)
        return ret
    elif signal.startswith('ana'):
        ret = analysisUsed(used)
        return ret


def analysisUsed(used):
    for u in used:
        if u == 'MakroSPS':
            used.remove('MakroSPS')
        elif u == 'MakroStep':
            used.remove('MakroStep')
        elif u == 'MakroTrigger':
            used.remove('MakroTrigger')
        elif u == 'Makro0':
            used.remove('Makro0')
        elif u == 'Makro20':
            used.remove('Makro20')
        if len(used) == 0:
            return False
    return True


def dealused(list):
    used = []
    flag = 'N'
    for i in list:
        if 'Folge' in i:
            flag = 'F'
        elif 'Makro' in i:
            flag = 'M'
        elif 'UP' in i:
            flag = 'U'
        if flag == 'F':
            if 'Folge' in i:
                used.append(i)
            else:
                used.append('Folge' + i)
        elif flag == 'M':
            if 'Makro' in i:
                used.append(i)
            else:
                used.append('Makro' + i)
        elif flag == 'U':
            if 'UP' in i:
                used.append(i)
            else:
                used.append('UP' + i)
    return used


def rob_standard_comment():
    ret = {}
    wb = load_workbook(STANDARD_COMMENT)
    ws = wb['机器人初始注释']
    key = ''
    for i in range(0, ws.max_row):
        for j in range(0, 4):
            val = ws.cell(row=i + 1, column=j + 1).value
            if j == 0 and re.search(r"^(E|A|M|I|ana|bin|t|F|T|Makro|UP|cell)\d{0,4}$", val) != None:
                key = val
            if j != 0 and key != '':
                if val == None:
                    if key not in ret:
                        ret[key] = ['Reserviert']
                else:
                    if key in ret:
                        ret[key].append(val)
                    else:
                        ret[key] = [val]
    return ret


def main():
    print('========================获取标注注释--start========================')
    standard_comment = rob_standard_comment()
    print('========================获取标注注释--end========================')

    print('========================主程序--start========================')
    RobotInfo(standard_comment, GET_BIG_DATA_SIMPLE)
    print('========================主程序--end========================')


if __name__ == "__main__":
    main()
