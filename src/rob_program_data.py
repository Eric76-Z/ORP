# 获取机器人数据
import json
import os
import re
import zipfile
from difflib import SequenceMatcher
import xml.dom.minidom

from openpyxl import Workbook, load_workbook
from tqdm import tqdm

from src.common.rob_program_backup_setting import SH_LOG_TITLE
from src.common.rob_program_data_setting import RobProgramData, SH_ROB_COMMENT_OVERVIEW
from src.common.rob_program_data_setting import PATH_PROGRAM, \
    ROB_PROGARM_DATA_EXPORT_FOLDER_BASE, ROB_PROGARM_DATA_EXPORT_OVERVIEW
from src.common.utils import TimeStampToTime, getFileSize, logWrite, createFolder, createSheet, backSignalSort


# 1、读取机器人zip相关信息
def RobotInfo():
    createFolder(ROB_PROGARM_DATA_EXPORT_FOLDER_BASE)
    wb = Workbook(write_only=True)
    createSheet(wb=wb, sh_name='机器人备份数据导出总览', sh_index=1, sh_title=SH_ROB_COMMENT_OVERVIEW)
    createSheet(wb=wb, sh_name='log', sh_index=2, sh_title=SH_LOG_TITLE)
    ov_path = os.path.join(ROB_PROGARM_DATA_EXPORT_FOLDER_BASE, ROB_PROGARM_DATA_EXPORT_OVERVIEW)
    wb.save(ov_path)
    wb.close()
    wb = load_workbook(ov_path)
    rob_program_data_json = {}

    for root, dirs, files in os.walk(PATH_PROGRAM):
        i = 0
        # 显示进度条
        for name in tqdm(files):
            originpath = os.path.join(root, name)
            if zipfile.is_zipfile(originpath):

                rob_program_data = RobProgramData()
                # ================path==================
                rob_program_data.path['path_origin'] = originpath  # 原始路径
                # ================meta==================
                rob_program_data.meta['title'] = name  # 文件名
                rob_program_data.meta['format'] = 'zip'  # 文件后缀
                rob_program_data.meta['mtime'] = TimeStampToTime(
                    os.path.getmtime(originpath))  # 原数据中修改时间，可视为创建时间
                rob_program_data.meta['size'] = getFileSize(originpath)
                # ================data==================
                rob_program_data.data['filename'] = name
                rob_program_data.data['controllername'] = name.split('.zip')[0]  # eg.k2a3a131s460r04
                rob_program_data.data['workstationname'] = rob_program_data.data['controllername'][
                                                           -7:].upper()  # 截取的 eg.s460r04
                rob_program_data.data['localLv1'] = rob_program_data.localLv1()
                rob_program_data.data['localLv2'] = rob_program_data.localLv2()
                rob_program_data.data['localLv3'] = rob_program_data.localLv3()

                # 创建文件夹
                ROB_PROGARM_DATA_EXPORT_FOLDER = ROB_PROGARM_DATA_EXPORT_FOLDER_BASE + '\\' + '机器人备份数据池' + '\\' + \
                                                 rob_program_data.data[
                                                     'localLv1'] + '\\' + \
                                                 rob_program_data.data['localLv2'] + '\\' + \
                                                 rob_program_data.data[
                                                     'localLv3']
                createFolder(ROB_PROGARM_DATA_EXPORT_FOLDER)
                ROB_PROGARM_DATA_EXPORT_FILENAME = rob_program_data.data['controllername'] + '.xlsx'
                rpd_path = os.path.join(ROB_PROGARM_DATA_EXPORT_FOLDER, ROB_PROGARM_DATA_EXPORT_FILENAME)  # 保存路径
                # print('========================抽取写入机器人备份数据--start========================')
                getData(rob_program_data, wb=wb)
                # exportData(rob_program_data, wb=wb, path=rpd_path)
                # print('========================抽取写入机器人备份数据--end========================')
                # 机器人overview表格写入
                ws = wb['机器人备份数据导出总览']
                i = i + 1
                ws.append([i, 'PFH2B', rob_program_data.data['localLv1'],
                           rob_program_data.data['localLv2'],
                           rob_program_data.data['localLv3'], rob_program_data.data['workstationname'],
                           rob_program_data.meta['mtime']])
                # 添加入rob_program_data_json
                rob_program_data_json[rob_program_data.data['workstationname']] = rob_program_data

            else:
                msg = '备份可能损坏!!!源路径为：' + originpath + ';'
                logWrite(wb=wb, controllername=name, sort='警告', msg=msg)
        # 写入json文件
        with open('database/rob_program_data.json', "w", encoding='utf-8') as f:
            f.write(json.dumps(rob_program_data_json, default=lambda obj: obj.__dict__, indent=4,
                               ensure_ascii=False))
            f.close()
        wb.save(ov_path)
        wb.close()


def getData(rob_program_data, wb):
    try:
        filezip = zipfile.ZipFile(rob_program_data.path['path_origin'], "r")
        rob_program_data.zipData['total_files'] = len(filezip.namelist())
        try:
            for file in filezip.namelist():
                if file.split('/')[-1] == 'am.ini':
                    am_ini = filezip.open(file)
                    mystr = str(am_ini.read())
                    rob_program_data.filedata['am.ini']['version'] = \
                        re.findall(r'Version=(.+)\[', mystr)[0].split('\\r\\n')[0]
                    rob_program_data.filedata['am.ini']['tech_packs'] = re.findall(r'TechPacks=(.+)\|', mystr)[0]
                elif file.split('/')[-1] == 'RobotInfo.xml':
                    RobotInfoXml = filezip.open(file)
                    dom = xml.dom.minidom.parse(RobotInfoXml)
                    root = dom.documentElement
                    rob_program_data.filedata['RobotInfo.xml']['serial_number'] = \
                        root.getElementsByTagName('SerialNumber')[0].childNodes[
                            0].data
                    rob_type = root.getElementsByTagName('RobotType')[0].childNodes[
                        0].data
                    rob_program_data.filedata['RobotInfo.xml']['rob_type'] = rob_type
                    rob_program_data.analysedata['rob']['rob_type'] = rob_type
                    # 区分机器人为轻载还是重载机器人
                    load_key = re.search(r'^(#KR)\d{1,3}', rob_type).group()
                    load_key_num = int(load_key.split('#KR')[1])
                    if load_key_num >= 210 and load_key_num <= 270:
                        rob_program_data.analysedata['rob']['load_type'] = 'light_load'
                    elif load_key_num >= 280 and load_key_num <= 500:
                        rob_program_data.analysedata['rob']['load_type'] = 'heavy_load'
                    else:
                        rob_program_data.analysedata['rob']['load_type'] = 'other_type'
                    rob_program_data.filedata['RobotInfo.xml']['time_stamp'] = \
                        root.getElementsByTagName('MamesOffsets')[
                            0].getAttribute('Timestamp')
                elif ((file.split('/')[-1] == 'A1.xml' or file.split('/')[-1] == 'A4.xml') and file.split('/')[
                    -2] == 'NGAxis') or (file.split('/')[-1] == 'ecatms_config.xml' and file.split('/')[
                    -2] == 'Cabinet'):
                    if file.split('/')[-1] == 'A1.xml':
                        ksp_xml = filezip.open(file)
                        dom = xml.dom.minidom.parse(ksp_xml)
                        root = dom.documentElement
                        ksp2 = root.getElementsByTagName('ToolMotor')[0].getAttribute('ServoFile')
                        rob_program_data.filedata['NGAxis']['A1.xml']['ksp2'] = ksp2
                        ksp2 = re.search(r"KSP\d{1,3}", ksp2)[0]
                        # KSP >> KSP 3*??
                        ksp2_list = list(ksp2)
                        ksp2_list.insert(3, ' 3*')
                        ksp2 = ''.join(ksp2_list)
                        rob_program_data.analysedata['cabinet']['ServoModuleFct']['ksp2'] = ksp2
                    elif file.split('/')[-1] == 'A4.xml':
                        ksp_xml = filezip.open(file)
                        dom = xml.dom.minidom.parse(ksp_xml)
                        root = dom.documentElement
                        ksp1 = root.getElementsByTagName('ToolMotor')[0].getAttribute('ServoFile')
                        rob_program_data.filedata['NGAxis']['A4.xml']['ksp1'] = ksp1
                        ksp1 = re.search(r"KSP\d{1,3}", ksp1)[0]
                        # KSP >> KSP 3*??
                        ksp1_list = list(ksp1)
                        ksp1_list.insert(3, ' 3*')
                        ksp1 = ''.join(ksp1_list)
                        rob_program_data.analysedata['cabinet']['ServoModuleFct']['ksp1'] = ksp1
                    elif file.split('/')[-1] == 'ecatms_config.xml':
                        ecatms_config_xml = filezip.open(file)
                        dom = xml.dom.minidom.parse(ecatms_config_xml)
                        root = dom.documentElement
                        kpp = root.getElementsByTagName('Slave')[2]
                        model = kpp.getElementsByTagName('Name')[0].childNodes[0].data
                        revision_no = kpp.getElementsByTagName('RevisionNo')[0].childNodes[0].data

                        rob_program_data.filedata['ecatms_config.xml']['kpp']['model'] = model  # kpp型号
                        rob_program_data.filedata['ecatms_config.xml']['kpp'][
                            'revision_no'] = revision_no  # 版本号，内含kpp具体型号
                        # 正则
                        model = re.search(r"KPP\d{1,3}", model)[0]
                        if model == 'KPP1' and revision_no == '196618':
                            model = 'KPP 600-20 1*64'
                        elif model == 'KPP1' and revision_no == '196617':
                            model = 'KPP 600-20 1*40'
                        elif model == 'KPP0':
                            model = 'KPP 600-20'
                        rob_program_data.analysedata['cabinet']['ServoModuleFct']['kpp'] = model
                elif (file.split('/')[-1] == 'E1.xml' or file.split('/')[-1] == 'E2.xml') and file.split('/')[
                    -2] == 'SimuAxis':
                    Xml = filezip.open(file)
                    dom = xml.dom.minidom.parse(Xml)
                    root = dom.documentElement
                    if file.split('/')[-1] == 'E1.xml':
                        rob_program_data.filedata['SimuAxis']['E1.xml']['machine_name'] = \
                            root.getElementsByTagName('Machine')[0].getAttribute('Name')

                        if rob_program_data.filedata['SimuAxis']['E1.xml']['machine_name'].startswith('#KR'):
                            rob_program_data.analysedata['cabinet']['extra_axis']['other_extra_axis'] = \
                                rob_program_data.filedata['SimuAxis']['E1.xml'][
                                    'machine_name']
                        else:
                            rob_program_data.analysedata['cabinet']['extra_axis']['linear_units'] = \
                                rob_program_data.filedata['SimuAxis']['E1.xml'][
                                    'machine_name']
                    elif file.split('/')[-1] == 'E2.xml':
                        rob_program_data.filedata['SimuAxis']['E2.xml']['machine_name'] = \
                            root.getElementsByTagName('Machine')[0].getAttribute('Name')
                        if rob_program_data.filedata['SimuAxis']['E2.xml']['machine_name'].startswith('#KR'):
                            rob_program_data.analysedata['cabinet']['extra_axis']['other_extra_axis'] = \
                                rob_program_data.filedata['SimuAxis']['E2.xml'][
                                    'machine_name']
                        else:
                            rob_program_data.analysedata['cabinet']['extra_axis']['linear_units'] = \
                                rob_program_data.filedata['SimuAxis']['E2.xml'][
                                    'machine_name']
                if rob_program_data.filedata['SimuAxis']['E1.xml']['machine_name'] != 'null' or \
                        rob_program_data.filedata['SimuAxis']['E2.xml']['machine_name'] != 'null':
                    rob_program_data.analysedata['cabinet']['extra_axis']['is_extra_axis'] = 'true'
                else:
                    rob_program_data.analysedata['cabinet']['extra_axis']['is_extra_axis'] = 'false'
        except Exception as e:
            print(e)
        rob_program_data.zipData['state'] = '备份完好'  # zip文件完好
        filezip.close()
    except Exception as e:
        msg = '备份可能损坏!!!源路径为：' + rob_program_data.path['path_origin'] + ';' + '新路径为:' + rob_program_data.path[
            'path_new']
        logWrite(wb=wb, controllername=rob_program_data.data['controllername'], sort='警告', msg=msg)
        rob_program_data.zipData['state'] = '备份损坏'


def main():
    print('========================主程序--start========================')
    RobotInfo()
    print('========================主程序--end========================')


if __name__ == "__main__":
    main()
