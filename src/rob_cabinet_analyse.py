# 获取机器人数据
import json
import os
import re
import time
import zipfile
from difflib import SequenceMatcher

from tqdm import tqdm
from openpyxl import Workbook, load_workbook

from src.common.constant import robtype_cabinet_map
from src.common.rob_cabinet_analyse_setting import ROB_CABINET_ANALYSE_FOLDER, ROB_CABINET_ANALYSE_FILE, \
    SH_ROB_CABINET_ANALYSE
from src.common.rob_program_backup_setting import SH_LOG_TITLE
from src.common.rob_program_comment_setting import RobProgramComment, SH_ROB_COMMENT_ANALYSE, PATH_PROGRAM, \
    ROB_COMMENT_OVERVIEW_REPORT_BASE, ROB_COMMENT_OVERVIEW_REPORT, \
    SH_ROB_COMMENT_OVERVIEW, STANDARD_COMMENT, BID_DATA_TEMPLE, SH_BID_DATA_TEMPLE, SUM
from src.common.setting import PATH_BASE, time_now, GET_BIG_DATA_SIMPLE
from src.common.utils import TimeStampToTime, getFileSize, logWrite, createFolder, createSheet, backSignalSort, \
    process_bar


# 1、读取机器人zip相关信息
def RobotCabinet():
    createFolder(ROB_CABINET_ANALYSE_FOLDER)
    wb = Workbook(write_only=True)
    createSheet(wb=wb, sh_name='机器人柜箱数据', sh_index=1, sh_title=SH_ROB_CABINET_ANALYSE)
    createSheet(wb=wb, sh_name='log', sh_index=2, sh_title=SH_LOG_TITLE)
    path = os.path.join(ROB_CABINET_ANALYSE_FOLDER, ROB_CABINET_ANALYSE_FILE)
    wb.save(path)
    wb.close()
    wb = load_workbook(path)
    ws = wb['机器人柜箱数据']
    with open('database/rob_program_data.json', 'r', encoding='utf-8') as f:
        info_dict = json.load(f)
        i = 0
        for dict in info_dict:
            i = i + 1
            localLv1 = info_dict[dict]['data']['localLv1']
            localLv2 = info_dict[dict]['data']['localLv2']
            localLv3 = info_dict[dict]['data']['localLv3']
            rob_type = info_dict[dict]['filedata']['RobotInfo.xml']['rob_type']
            version = info_dict[dict]['filedata']['am.ini']['version']
            load_type = info_dict[dict]['analysedata']['rob']['load_type']
            if info_dict[dict]['analysedata']['cabinet']['extra_axis']['linear_units'] == 'null':
                is_linear_units = 'no_linear_units'
            else:
                is_linear_units = 'linear_units'
            if version == 'V8.3.24':
                continue
            print(dict)
            KSP2_SET = robtype_cabinet_map[version][load_type][is_linear_units]['KSP2']
            KSP1_SET = robtype_cabinet_map[version][load_type][is_linear_units]['KSP1']
            KPP_SET = robtype_cabinet_map[version][load_type][is_linear_units]['KPP']
            KSP2_ACT = info_dict[dict]['analysedata']['cabinet']['ServoModuleFct']['ksp2']
            KSP1_ACT = info_dict[dict]['analysedata']['cabinet']['ServoModuleFct']['ksp1']
            KPP_ACT = info_dict[dict]['analysedata']['cabinet']['ServoModuleFct']['kpp']

            if KSP2_ACT == KSP2_SET and KSP1_ACT == KSP1_SET and KPP_ACT == KPP_SET:
                cabinet_ok = 'true'
            else:
                cabinet_ok = 'false'
            content = [i + 1, localLv1, localLv2, localLv3,
                       dict, rob_type, KSP2_SET, KSP1_SET, KPP_SET, KSP2_ACT, KSP1_ACT, KPP_ACT, cabinet_ok
                       ]
            ws.append(content)

        wb.save(path)
        wb.close()


def main():
    print('========================主程序--start========================')
    RobotCabinet()
    print('========================主程序--end========================')


if __name__ == "__main__":
    main()
