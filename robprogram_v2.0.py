import json
import os
import shutil
import time
import xml.dom.minidom
import zipfile

import xlrd
from openpyxl import Workbook, load_workbook
from xlutils.copy import copy

from ini import RobProgramData, PATH_ORIGIN_BACKUP, PATH_EXPORT_TO, PATH_TRASH, LOG_TARSH_NAME, PATH_BASE, \
    LOG_FILE_NAME, PATH_REPORT, FILE_STANDARD, PATH_STANDARD, FILE_REPORT, SH_LOG_TITLE, SH_ROB_BACKUP_OV, \
    SH_ROB_BACKUP_ANA
from pathmap import pathmap

# 获取RobotProgram对象信息，并写入
from utils import TimeStampToTime, logWrite, logWriteTitle, getFileSize, createFolder, createSheet


# 获取机器人数据
def RobotInfo(SUM, wb):
    rob_program_data_json = {}
    for root, dirs, files in os.walk(PATH_ORIGIN_BACKUP):
        for name in files:
            if name.endswith('.zip'):
                rob_program_data = RobProgramData()
                SUM['total_files'] = SUM['total_files'] + 1
                originpath = os.path.join(root, name)
                # ================path==================
                rob_program_data.path['path_origin'] = originpath  # 原始路径
                # ================meta==================
                rob_program_data.meta['title'] = name  # 文件名
                rob_program_data.meta['format'] = 'zip'  # 文件后缀
                rob_program_data.meta['mtime'] = TimeStampToTime(os.path.getmtime(originpath))  # 原数据中修改时间，可视为创建时间
                rob_program_data.meta['size'] = getFileSize(originpath)
                # ================data==================
                rob_program_data.data['filename'] = name
                rob_program_data.data['controllername'] = name.split('.zip')[0]  # eg.k2a3a131s460r04
                rob_program_data.data['workstationname'] = rob_program_data.data['controllername'][
                                                           -7:].upper()  # 截取的 eg.s460r04
                rob_program_data.data['localLv1'] = rob_program_data.localLv1()
                rob_program_data.data['localLv2'] = rob_program_data.localLv2()
                rob_program_data.data['localLv3'] = rob_program_data.localLv3()

                # ================error==================
                if rob_program_data.data['localLv1'] == '':
                    logWrite(wb=wb, controllername=rob_program_data.data['controllername'], sort='报错', msg='没有对应的一级地点')
                    SUM['err_files'] = SUM['err_files'] + 1
                    continue
                elif rob_program_data.data['localLv2'] == '':
                    logWrite(wb=wb, controllername=rob_program_data.data['controllername'], sort='报错', msg='没有对应二级地点')
                    SUM['err_files'] = SUM['err_files'] + 1
                    continue
                elif rob_program_data.data['localLv3'] == '':
                    logWrite(wb=wb, controllername=rob_program_data.data['controllername'], sort='报错', msg='没有对应三级地点')
                    SUM['err_files'] = SUM['err_files'] + 1
                    continue
                rob_program_data.path['path_new'] = PATH_EXPORT_TO + '\\' + rob_program_data.data['localLv1'] + '\\' + \
                                                    rob_program_data.data['localLv2'] + '\\' + rob_program_data.data[
                                                        'localLv3'] + '\\' + name

                # print(int(rob_program_data.meta['size'].split('.')[0]))
                # 判断rob_program_data中是否已有这个工位
                if rob_program_data.data['workstationname'] in rob_program_data_json.keys():
                    SUM['repeat_files'] = SUM['repeat_files'] + 1
                    # 进一步判断两者创建时间
                    if rob_program_data.meta['mtime'] == \
                            rob_program_data_json[rob_program_data.data['workstationname']].meta['mtime']:
                        if rob_program_data.meta['size'] == \
                                rob_program_data_json[rob_program_data.data['workstationname']].meta['size']:
                            # 如果文件创建时间一致，大小一致，则将待处理文件移动到删除区
                            shutil.move(rob_program_data.path['path_origin'], PATH_TRASH)
                            logWrite(wb=wb, controllername=rob_program_data.data['controllername'], sort='警告',
                                     msg='重复文件，创建时间一致，大小一致，已将其中一个备份移动至垃圾箱！')
                            SUM['trash_files'] = SUM['trash_files'] + 1
                            continue
                        else:
                            # 如果文件创建时间一致，大小不一致，谁小删谁
                            if int(rob_program_data.meta['size'].split('.')[0]) >= \
                                    int(rob_program_data_json[rob_program_data.data['workstationname']].meta[
                                            'size'].split('.')[0]):
                                shutil.move(rob_program_data_json[rob_program_data.data['workstationname']].path[
                                                'path_origin'], PATH_TRASH)
                                SUM['trash_files'] = SUM['trash_files'] + 1
                            else:
                                shutil.move(rob_program_data.path['path_origin'], PATH_TRASH)
                                SUM['delete_files'] = SUM['delete_files'] + 1
                            logWrite(wb=wb, controllername=rob_program_data.data['controllername'], sort='警告',
                                     msg='重复文件，创建时间一致，大小不一致，且size小的备份被移动至垃圾箱！')
                            continue
                    else:
                        # 备份时间判断
                        if rob_program_data.meta['mtime'] > \
                                rob_program_data_json[rob_program_data.data['workstationname']].meta['mtime']:
                            try:
                                shutil.move(rob_program_data_json[rob_program_data.data['workstationname']].path[
                                                'path_origin'], PATH_TRASH)
                                SUM['trash_files'] = SUM['trash_files'] + 1
                            except:
                                os.remove(rob_program_data_json[rob_program_data.data['workstationname']].path[
                                              'path_origin'])
                        else:
                            try:
                                shutil.move(rob_program_data.path['path_origin'], PATH_TRASH)
                            except:
                                os.remove(rob_program_data.path['path_origin'])
                        logWrite(wb=wb, controllername=rob_program_data.data['controllername'], sort='警告',
                                 msg='备份的创建时间过早，移动至垃圾箱！')
                        continue
                else:
                    pass
                # 分析备份大小
                SUM['avg_file_size'] = (SUM['avg_file_size'] * len(rob_program_data_json) + float(
                    rob_program_data.meta['size'].split(' ')[0])) / (len(
                    rob_program_data_json) + 1)
                SUM['avg_file_size'] = float(format(SUM['avg_file_size'], '.2f'))
                if float(rob_program_data.meta['size'].split(' ')[0]) > float(SUM['max_file_size'].split(' ')[0]):
                    SUM['max_file_size'] = rob_program_data.meta['size']
                elif float(rob_program_data.meta['size'].split(' ')[0]) < float(SUM['min_file_size'].split(' ')[0]):
                    SUM['min_file_size'] = rob_program_data.meta['size']
                if float(rob_program_data.meta['size'].split(' ')[0]) > 40 or float(
                        rob_program_data.meta['size'].split(' ')[0]) < 10:
                    msg = '备份的size大于40M或者小于10M！为：' + rob_program_data.meta['size']
                    logWrite(wb=wb, controllername=rob_program_data.data['controllername'], sort='提示', msg=msg)
                    continue
                # print('========================解析机器人备份--start========================')
                analysisZip(rob_program_data, wb=wb)
                # print('========================解析机器人备份--end========================')
                # 添加入rob_program_data_json

                rob_program_data_json[rob_program_data.data['workstationname']] = rob_program_data
    with open('robot_data_json.json', "w", encoding='utf-8') as f:
        f.write(json.dumps(rob_program_data_json, default=lambda obj: obj.__dict__, indent=4, ensure_ascii=False))
        f.close()


def backupOverview(wb, SUM):
    ## 读取json文件
    book_standard = xlrd.open_workbook(os.path.join(PATH_STANDARD, FILE_STANDARD), formatting_info=True)
    sh_standard = book_standard.sheet_by_name('RobStandard')
    nrows = sh_standard.nrows
    sh = wb['机器人备份总览']
    # content = []
    with open('robot_data_json.json', 'r', encoding='utf-8') as f:
        info_dict = json.load(f)
        for i in range(1, nrows):
            # print(sh_standard.cell_value(i, 5))
            depart = sh_standard.cell_value(i, 1)
            localLv1 = sh_standard.cell_value(i, 2)
            localLv2 = sh_standard.cell_value(i, 3)
            localLv3 = sh_standard.cell_value(i, 4)
            create_time = ''
            size = ''
            isOK = ''
            totalFiles = 0
            if sh_standard.cell_value(i, 5) in info_dict:
                localLv1 = info_dict[sh_standard.cell_value(i, 5)]['data']['localLv1']
                localLv2 = info_dict[sh_standard.cell_value(i, 5)]['data']['localLv2']
                localLv3 = info_dict[sh_standard.cell_value(i, 5)]['data']['localLv3']
                create_time = info_dict[sh_standard.cell_value(i, 5)]['meta']['mtime']
                size = info_dict[sh_standard.cell_value(i, 5)]['meta']['size']
                isOK = info_dict[sh_standard.cell_value(i, 5)]['zipData']['isOK']
                totalFiles = info_dict[sh_standard.cell_value(i, 5)]['zipData']['total_files']
            content = [sh_standard.cell_value(i, 0), depart, localLv1, localLv2, localLv3, sh_standard.cell_value(i, 5),
                       create_time, size, isOK, totalFiles]
            sh.append(content)


def Reforming(SUM):
    ## 移动重整文件
    ## 读取json文件
    with open('robot_data_json.json', 'r', encoding='utf-8') as f:
        info_dict = json.load(f)
        for dict in info_dict:
            folder = os.path.exists(info_dict[dict]['path']['path_new'])
            if not folder:  # 判断是否存在文件夹如果不存在则创建为文件夹
                os.makedirs(info_dict[dict]['path']['path_new'])  # makedirs 创建文件时如果路径不存在会创建这个路径
            else:
                pass
            if not os.path.exists(info_dict[dict]['path']['path_new']):
                SUM['move_files'] = SUM['move_files'] + 1
                shutil.copy2(info_dict[dict]['path']['path_origin'], info_dict[dict]['path']['path_new'])
            else:
                SUM['exists_files'] = SUM['exists_files'] + 1
                continue
    print(SUM)


def analysisZip(rob_program_data, wb):
    try:
        filezip = zipfile.ZipFile(rob_program_data.path['path_origin'], "r")
        rob_program_data.zipData['total_files'] = len(filezip.namelist())
        try:
            for file in filezip.namelist():
                if (file.split('/')[-1].endswith('.src')):
                    if (file.split('/')[-1].startswith('Folge')):
                        rob_program_data.zipData['file_folge_num'] = rob_program_data.zipData['file_folge_num'] + 1
                    elif (file.split('/')[-1].startswith('makro')):
                        rob_program_data.zipData['file_makro_num'] = rob_program_data.zipData['file_makro_num'] + 1
                    elif (file.split('/')[-1].startswith('up')):
                        rob_program_data.zipData['up'] = rob_program_data.zipData['up'] + 1
                elif file.split('/')[-1] == 'RobotInfo.xml':
                    dom = xml.dom.minidom.parse(file)
                    root = dom.documentElement
                    print(root.nodeName)
            filezip.close()
            rob_program_data.zipData['isOK'] = 'OK'
        except Exception as e:
            print(rob_program_data.path['path_origin'] + str(e))

    except Exception as e:
        msg = '备份可能损坏!!!源路径为：' + rob_program_data.path['path_origin'] + ';' + '新路径为:' + rob_program_data.path[
            'path_new']
        logWrite(wb=wb, controllername=rob_program_data.data['controllername'], sort='警告', msg=msg)
        rob_program_data.zipData['isOK'] = 'NOK'


def backupState():
    standard_filepath = buStandard['filepath'] + '\\' + buStandard['filename']
    book_rd = xlrd.open_workbook(standard_filepath, formatting_info=True)
    sheet_rd = book_rd.sheet_by_index(0)
    book_wt = copy(book_rd)
    sheet_wt = book_wt.get_sheet(SHEET_NAME)
    nrows = sheet_rd.nrows
    nrows_compare = nrows
    global DEAL_FILES
    for root, dirs, files in os.walk(PATH_EXPORT_TO):
        for name in files:
            DEAL_FILES = DEAL_FILES + 1
            rob_program_data.controllername = name.split('.zip')[0]
            workstationname = rob_program_data.controllername[-7:].upper()  # 截取的 eg.k2a3a131s460r04 后7位
            if isAll == True:
                compare_name = rob_program_data.controllername
            else:
                compare_name = workstationname
            flag = True

            for i in range(nrows):
                lng = sheet_rd.cell(i - 1, COMPARE_COL).value
                if lng == compare_name:
                    sheet_wt.write(i - 1, COMMIT_COL, label='已备份')
                    sheet_wt.write(i - 1, TIME_COL, label=time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
                    flag = False
                    break
            if flag == True:
                sheet_wt.write(nrows_compare, COMPARE_COL - 3, targetData[compare_name]['Lv1'])
                sheet_wt.write(nrows_compare, COMPARE_COL - 2,
                               pathmap[rob_program_data.controllername[2:6].lower()]['Lv2'])
                sheet_wt.write(nrows_compare, COMPARE_COL - 1,
                               pathmap[rob_program_data.controllername[2:6].lower()]['Lv3'])
                sheet_wt.write(nrows_compare, COMPARE_COL, compare_name)
                sheet_wt.write(nrows_compare, COMMIT_COL, label='新工位')
                sheet_wt.write(nrows_compare, TIME_COL, label=time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
                nrows_compare = nrows_compare + 1
    logWriteTitle('总结')
    log = open(PATH_BASE + '\\' + LOG_FILE_NAME, 'a')
    log.write(
        '备份总数: ' + str(SUM['total_files']) + '        已处理: ' + str(DEAL_FILES) + '         异常: ' + str(
            SUM['err_files']) + '\r\n')
    log.close()
    logWriteTitle('end')
    book_wt.save(PATH_BASE + '\\' + time.strftime("%Y%m%d", time.localtime()) + '机器人备份情况.xls')


def main():
    # 总结
    global SUM
    SUM = {
        'total_files': 0,  # 更目录下文件总数
        'err_files': 0,
        'move_files': 0,
        'exists_files': 0,
        'repeat_files': 0,
        'trash_files': 0,
        'avg_file_size': 0,
        'min_file_size': '1000.0 MB',
        'max_file_size': '00.0 MB'
    }
    createFolder(PATH_TRASH)
    createFolder(PATH_REPORT)
    # logWriteTitle(PATH_BASE + '\\' + LOG_FILE_NAME, 'start')
    # logWriteTitle(PATH_TRASH + '\\' + LOG_TARSH_NAME, 'start')
    # wb =
    wb = Workbook(write_only=True)
    createSheet(wb=wb, sh_name='机器人备份总览', sh_index=1, sh_title=SH_ROB_BACKUP_OV)
    createSheet(wb=wb, sh_name='机器人备份分析', sh_index=2, sh_title=SH_ROB_BACKUP_ANA)
    createSheet(wb=wb, sh_name='log', sh_index=3, sh_title=SH_LOG_TITLE)

    print('========================获取机器人数据--start========================')
    RobotInfo(SUM, wb)
    print('========================获取机器人数据--end========================')
    print('========================移动重整备份--start========================')
    Reforming(SUM)
    print('========================移动重整备份--end========================')
    print('========================机器人overview--start========================')
    backupOverview(wb, SUM)
    print('========================机器人overview--end========================')
    #
    # logWriteTitle(PATH_BASE + '\\' + LOG_FILE_NAME, 'end')
    # logWriteTitle(PATH_TRASH + '\\' + LOG_TARSH_NAME, 'end')
    #
    # backupState()

    # # 2.解压某文件到指定文件夹
    # extractFile()
    #
    # # 3.解析文件，提取数据
    # getfileData()
    #
    # # 4.导出数据
    # exportfileData()
    wb.save(os.path.join(PATH_REPORT, FILE_REPORT))


if __name__ == "__main__":
    main()
